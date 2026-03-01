"""
src/exhibit_builder.py
=======================
Converts a normalised long-format DataFrame into a styled Excel exhibit.
Reuses the same openpyxl patterns as agentic-claims-analyst for consistency.
"""

from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NAVY  = "1F3864"
BLUE  = "2E75B6"
GREY  = "D6E4F0"
WHITE = "FFFFFF"


def build_exhibit_from_df(df: pd.DataFrame, output_path: str,
                          source_pdf: str = "") -> str:
    """
    Build a formatted Excel exhibit workbook from a normalised DataFrame.
    Returns the resolved output path.
    """
    if df.empty:
        # Create a blank workbook with a note
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = f"No data extracted from: {source_pdf}"
        wb.save(output_path)
        return output_path

    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, df, source_pdf)
    _build_pmpm_summary(wb, df)
    _build_detail_sheet(wb, df)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    path = str(Path(output_path).resolve())
    wb.save(path)
    return path


def _build_cover_sheet(wb, df: pd.DataFrame, source_pdf: str):
    ws = wb.create_sheet("Cover")

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Medical Claims Analysis — Extracted from PDF"
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    details = [
        ("Source PDF",   source_pdf or "N/A"),
        ("Extracted",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Records",      f"{len(df):,}"),
        ("Date Range",   f"{df['month'].min().strftime('%b %Y')} – "
                         f"{df['month'].max().strftime('%b %Y')}"),
        ("LOBs",         ", ".join(df["lob"].dropna().unique())),
        ("Providers",    ", ".join(df["provider"].dropna().unique())),
        ("Total Members (avg/mo)",
         f"{df.groupby('month')['members'].sum().mean():,.0f}" if "members" in df.columns else "N/A"),
    ]

    for r, (label, value) in enumerate(details, 3):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=GREY)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45


def _build_pmpm_summary(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Exhibit 1 - PMPM Summary")

    _header_block(ws, "PMPM Trend Summary — All Lines of Business")

    providers = df["provider"].dropna().unique().tolist()
    months    = sorted(df["month"].unique())

    # Build pivot: weighted avg PMPM per provider per month
    pivot_rows = []
    for month in months:
        sub = df[df["month"] == month]
        row = {"Month": pd.Timestamp(month).strftime("%b %Y")}
        for p in providers:
            psub = sub[sub["provider"] == p]
            if not psub.empty and psub["members"].sum() > 0:
                wpmpm = (psub["pmpm"] * psub["members"]).sum() / psub["members"].sum()
            else:
                wpmpm = None
            row[p] = round(wpmpm, 2) if wpmpm else None
        pivot_rows.append(row)

    start = 5
    headers = ["Month"] + providers
    _write_header(ws, start, 1, headers)

    for r, row in enumerate(pivot_rows, start + 1):
        ws.cell(r, 1, row["Month"]).alignment = Alignment(horizontal="center")
        for c, p in enumerate(providers, 2):
            cell = ws.cell(r, c, row.get(p))
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)

    _set_widths(ws, {1: 14, 2: 14, 3: 14, 4: 14})

    # Line chart
    end_row = start + len(pivot_rows)
    chart = LineChart()
    chart.title   = "PMPM Trend by Provider"
    chart.y_axis.title = "PMPM ($)"
    chart.height  = 12
    chart.width   = 22
    chart.style   = 10

    # Add all provider series at once using add_data
    data = Reference(ws, min_col=2, max_col=1 + len(providers),
                     min_row=start, max_row=start + len(pivot_rows))
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=start + 1, max_row=start + len(pivot_rows))
    chart.set_categories(cats)
    ws.add_chart(chart, f"F{start}")


def _build_detail_sheet(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Exhibit 2 - LOB Detail")
    _header_block(ws, "Detailed PMPM by Line of Business & Provider")

    lobs      = df["lob"].dropna().unique().tolist()
    providers = df["provider"].dropna().unique().tolist()
    months    = sorted(df["month"].unique())

    # Write sub-tables per LOB stacked vertically
    current_row = 5
    for lob in lobs:
        sub = df[df["lob"] == lob]

        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=len(providers) + 2
        )
        c = ws.cell(current_row, 1, lob)
        c.font  = Font(bold=True, color=WHITE, size=11)
        c.fill  = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center")
        current_row += 1

        _write_header(ws, current_row, 1, ["Month", "Members"] + providers)
        current_row += 1

        for month in months:
            msub = sub[sub["month"] == month]
            ws.cell(current_row, 1, pd.Timestamp(month).strftime("%b %Y"))
            ws.cell(current_row, 2, int(msub["members"].sum()) if not msub.empty else None)
            ws.cell(current_row, 2).number_format = "#,##0"
            for c, p in enumerate(providers, 3):
                psub = msub[msub["provider"] == p]
                val  = round(psub["pmpm"].values[0], 2) if not psub.empty and psub["pmpm"].notna().any() else None
                cell = ws.cell(current_row, c, val)
                cell.number_format = '"$"#,##0.00'
                if current_row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=GREY)
            current_row += 1

        current_row += 2  # spacing between LOB blocks

    _set_widths(ws, {i: 14 for i in range(1, 8)})


# ── Helpers ────────────────────────────────────────────────────────────────────
def _header_block(ws, title: str):
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = title
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24


def _write_header(ws, row: int, start_col: int, headers: list):
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(row, c, h)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")


def _set_widths(ws, widths: dict):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
